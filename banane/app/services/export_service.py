# app/services/export_service.py
# ============================================================
#  UBUNTUTECH — Service Export PDF + Excel v2.0
#  FIX Railway : génération en mémoire (BytesIO), pas sur disque
# ============================================================
import io
from datetime import datetime, timedelta
from typing import Optional
from sqlalchemy.orm import Session
from sqlalchemy import func
from loguru import logger

from app.models.utilisateur import Utilisateur
from app.models.boutique import Boutique
from app.models.vente import Vente, LigneVente
from app.models.depense import Depense
from app.models.produit import Produit
from app.models.client import Client


class ExportService:

    # ── Bilan PDF — retourne bytes en mémoire ─────────────────
    @staticmethod
    def bilan_pdf_bytes(
        db: Session,
        user: Utilisateur,
        id_boutique: int,
        periode: str = "mois",
    ) -> bytes:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        boutique = db.query(Boutique).filter(Boutique.id_boutique == id_boutique).first()
        if not boutique:
            raise ValueError("Boutique introuvable")

        now = datetime.utcnow()
        if periode == "mois":
            debut = now.replace(day=1, hour=0, minute=0, second=0)
            label_periode = now.strftime("%B %Y")
        elif periode == "semaine":
            debut = now - timedelta(days=7)
            label_periode = "7 derniers jours"
        elif periode == "jour":
            debut = now.replace(hour=0, minute=0, second=0, microsecond=0)
            label_periode = "Aujourd'hui"
        elif periode == "trimestre":
            debut = now - timedelta(days=90)
            label_periode = "3 derniers mois"
        elif periode == "annee":
            debut = now.replace(month=1, day=1, hour=0, minute=0, second=0)
            label_periode = str(now.year)
        else:
            debut = now - timedelta(days=30)
            label_periode = "30 derniers jours"

        revenus = float(db.query(func.sum(Vente.montant_total)).filter(
            Vente.id_boutique == id_boutique,
            Vente.statut == "validee",
            Vente.date_vente >= debut,
        ).scalar() or 0)

        nb_ventes = db.query(func.count(Vente.id_vente)).filter(
            Vente.id_boutique == id_boutique,
            Vente.statut == "validee",
            Vente.date_vente >= debut,
        ).scalar() or 0

        depenses = float(db.query(func.sum(Depense.montant)).filter(
            Depense.id_boutique == id_boutique,
            Depense.date_depense >= debut,
        ).scalar() or 0)

        benefice = revenus - depenses

        top_produits = db.query(
            Produit.nom_produit,
            func.sum(LigneVente.montant_ligne).label("total"),
        ).join(LigneVente, LigneVente.id_produit == Produit.id_produit
        ).join(Vente, Vente.id_vente == LigneVente.id_vente
        ).filter(
            Vente.id_boutique == id_boutique,
            Vente.statut == "validee",
            Vente.date_vente >= debut,
        ).group_by(Produit.nom_produit).order_by(func.sum(LigneVente.montant_ligne).desc()).limit(10).all()

        dernieres_ventes = db.query(Vente).filter(
            Vente.id_boutique == id_boutique,
            Vente.statut == "validee",
            Vente.date_vente >= debut,
        ).order_by(Vente.date_vente.desc()).limit(15).all()

        # Générer en mémoire
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        vert   = colors.HexColor("#1B8A4C")
        orange = colors.HexColor("#F47B20")

        titre_style = ParagraphStyle("titre", parent=styles["Title"],  textColor=vert,  fontSize=18, spaceAfter=6)
        sous_style  = ParagraphStyle("sous",  parent=styles["Normal"], textColor=colors.gray, fontSize=10, spaceAfter=4)
        h2_style    = ParagraphStyle("h2",    parent=styles["Heading2"], textColor=vert, fontSize=12, spaceBefore=12, spaceAfter=6)
        footer_style = ParagraphStyle("footer", parent=styles["Normal"], fontSize=8, textColor=colors.gray, alignment=TA_CENTER)

        elements = [
            Paragraph("UBUNTUTECH", titre_style),
            Paragraph(f"Bilan Financier — {boutique.nom_boutique}", sous_style),
            Paragraph(f"Période : {label_periode}  |  Généré le {now.strftime('%d/%m/%Y à %H:%M')}", sous_style),
            Paragraph(f"Commerçant : {user.nom_complet}  |  Score santé : {user.score_sante_business}/100", sous_style),
            HRFlowable(color=vert, thickness=2, spaceAfter=12),
            Paragraph("Résumé Financier", h2_style),
        ]

        resume_data = [
            ["Indicateur", "Valeur", ""],
            ["Chiffre d'affaires", f"{revenus:,.0f} FCFA", ""],
            ["Total dépenses",     f"{depenses:,.0f} FCFA", ""],
            ["Bénéfice net",       f"{benefice:,.0f} FCFA", "✅" if benefice >= 0 else "⚠️"],
            ["Nombre de ventes",   str(nb_ventes), ""],
            ["Panier moyen",       f"{revenus/nb_ventes:,.0f} FCFA" if nb_ventes else "N/A", ""],
        ]
        t = Table(resume_data, colWidths=[7*cm, 5*cm, 2*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), vert),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 10),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#F5FAF7")]),
            ("GRID",          (0,0), (-1,-1), 0.5, colors.lightgrey),
            ("TEXTCOLOR",     (1,3), (1,3),   vert if benefice >= 0 else colors.red),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 0.5*cm))

        if top_produits:
            elements.append(Paragraph("Top Produits", h2_style))
            prod_data = [["Produit", "CA Généré"]] + [
                [p.nom_produit, f"{float(p.total):,.0f} FCFA"] for p in top_produits
            ]
            tp = Table(prod_data, colWidths=[10*cm, 4*cm])
            tp.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,0), orange),
                ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
                ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",      (0,0), (-1,-1), 9),
                ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#FFF8F3")]),
                ("GRID",          (0,0), (-1,-1), 0.5, colors.lightgrey),
                ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                ("TOPPADDING",    (0,0), (-1,-1), 5),
            ]))
            elements.append(tp)
            elements.append(Spacer(1, 0.5*cm))

        if dernieres_ventes:
            elements.append(Paragraph("Dernières Ventes", h2_style))
            ventes_data = [["Date", "ID", "Montant", "Mode"]] + [
                [v.date_vente.strftime("%d/%m %H:%M"), f"#{v.id_vente}", f"{float(v.montant_total):,.0f} F", v.mode_paiement]
                for v in dernieres_ventes
            ]
            tv = Table(ventes_data, colWidths=[3*cm, 4*cm, 4*cm, 3*cm])
            tv.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,0), vert),
                ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
                ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",      (0,0), (-1,-1), 8),
                ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#F5FAF7")]),
                ("GRID",          (0,0), (-1,-1), 0.5, colors.lightgrey),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                ("TOPPADDING",    (0,0), (-1,-1), 4),
            ]))
            elements.append(tv)

        elements.append(Spacer(1, cm))
        elements.append(HRFlowable(color=colors.lightgrey, thickness=1))
        elements.append(Paragraph(
            f"UbuntuTech v2.0 — «Je réussis parce que nous réussissons» — {boutique.ville}",
            footer_style
        ))

        doc.build(elements)
        buffer.seek(0)
        logger.info(f"Bilan PDF généré en mémoire pour boutique {id_boutique}")
        return buffer.getvalue()

    # ── Stock Excel — retourne bytes en mémoire ───────────────
    @staticmethod
    def stock_excel_bytes(db: Session, id_boutique: int) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        produits = db.query(Produit).filter(
            Produit.id_boutique == id_boutique,
            Produit.actif == True
        ).order_by(Produit.nom_produit).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventaire"

        vert      = "1B8A4C"
        vert_clair = "E8F5EF"
        orange    = "F47B20"

        headers = ["Produit", "Unité", "Qté Stock", "Seuil Alerte", "Statut",
                   "Prix Vente", "Prix Achat", "Marge %", "Nb Ventes", "Total Vendu"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor=vert)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

        for row, p in enumerate(produits, 2):
            qte   = float(p.quantite_stock or 0)
            seuil = float(p.seuil_alerte_stock or 5)
            marge = float(p.marge_pct or 0) if p.marge_pct else (
                ((float(p.prix_vente) - float(p.prix_achat)) / float(p.prix_vente) * 100)
                if float(p.prix_vente) > 0 else 0
            )
            if qte <= 0:
                statut = "RUPTURE"
                fill   = PatternFill("solid", fgColor="FFCCCC")
            elif qte <= seuil:
                statut = "ALERTE"
                fill   = PatternFill("solid", fgColor="FFE5CC")
            else:
                statut = "OK"
                fill   = PatternFill("solid", fgColor=vert_clair) if row % 2 == 0 else None

            for col, val in enumerate([
                p.nom_produit, p.unite, qte, seuil, statut,
                float(p.prix_vente), float(p.prix_achat),
                round(marge, 1), int(p.nb_ventes or 0), float(p.total_vendu or 0)
            ], 1):
                cell = ws.cell(row=row, column=col, value=val)
                if fill:
                    cell.fill = fill

        for i, w in enumerate([25, 8, 10, 10, 10, 12, 12, 10, 10, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(f"Stock Excel généré en mémoire pour boutique {id_boutique}")
        return buffer.getvalue()

    # ── Crédits clients PDF — retourne bytes en mémoire ───────
    @staticmethod
    def credits_pdf_bytes(db: Session, id_boutique: int) -> bytes:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        clients_credit = db.query(Client).filter(
            Client.id_boutique == id_boutique,
            Client.solde_credit > 0,
            Client.actif == True
        ).order_by(Client.solde_credit.desc()).all()

        total_credit = sum(float(c.solde_credit) for c in clients_credit)
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        vert   = colors.HexColor("#1B8A4C")

        elements = [
            Paragraph("UBUNTUTECH — Suivi Crédits Clients", ParagraphStyle("t", parent=styles["Title"], textColor=vert, fontSize=14)),
            Paragraph(f"Total crédits en attente : {total_credit:,.0f} FCFA — {len(clients_credit)} client(s)", styles["Normal"]),
            Spacer(1, 0.5*cm),
        ]

        data = [["Client", "Téléphone", "Quartier", "Crédit (FCFA)", "Fiabilité", "Dernier achat"]]
        for c in clients_credit:
            data.append([
                c.nom_client,
                c.telephone or "",
                getattr(c, "quartier", "") or "",
                f"{float(c.solde_credit):,.0f}",
                f"{c.fiabilite_paiement}/100",
                c.derniere_visite.strftime("%d/%m/%Y") if c.derniere_visite else "",
            ])

        t = Table(data, colWidths=[5*cm, 3*cm, 3*cm, 3*cm, 2.5*cm, 2.5*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), vert),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#F5FAF7")]),
            ("GRID",          (0,0), (-1,-1), 0.5, colors.lightgrey),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]))
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
